# excel test：name is:pythonexcel1.py

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print(sheet.max_row)
print(sheet.max_column)
