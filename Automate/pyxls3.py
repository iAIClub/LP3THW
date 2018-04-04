import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_active_sheet()
# list(sheet.columns)[1]
for cell in list(sheet.columns)[1]:
    print(cell.value)
'''
#for cellObj in sheet.columns[1]:
#        print(cellObj.value)
# print columns[1]
for column in sheet.columns:
  for cell in column:
    print(cell.value）
'''

'''
for column in sheet.columns:
    for cell in column:
        print(cell.value)
'''



'''
  a)逐行读
        ws.iter_rows(range_string=None, row_offset=0, column_offset=0): range-string(string)-单元格的范围：例如('A1:C4') row_offset-添加行 column_offset-添加列
    返回一个生成器, 注意取值时要用value,例如：
    for row in ws.iter_rows('A1:C2'):
        for cell in row:
            print cell
   读指定行、指定列:
    rows=ws.rows#row是可迭代的
    columns=ws.columns#column是可迭代的
    打印第n行数据
    print rows[n]#不需要用.value
    print columns[n]#不需要用.value
'''
